"""
rebuild_recurring_v2.py

Reads the existing Recurring tab (already rebuilt by v1) and applies
Eric Carlson's feedback from 02/25/2026:

1. Include history months: Nov, Dec, Jan (+ current + forward)
1. Month columns = reporting month, not distribution month
1. Remove "Timing" column (same as GV Day — redundant)
1. Quarterly items: leave non-quarter-end months blank
1. Replace "Katie" with "Brian" everywhere (owner/preparer)
1. Remove "AIPI" from the deliverables list

Usage:
    pip install openpyxl
    python rebuild_recurring_v2.py
"""

import re
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────────
FILE_PATH = r"C:\Users\iams395\Copy of Project Tracking.xlsx"
TAB_NAME = "Recurring"

# How many history months to include before the current month
HISTORY_MONTHS = 3   # Nov, Dec, Jan
FORWARD_MONTHS = 5   # Mar through Jul (current month = Feb is included too)

# ── RBC COLOR PALETTE ────────────────────────────────────────────────────────
RBC_NAVY = "001F5B"
RBC_BLUE = "0051A5"
RBC_LIGHT_BLUE = "C8D9F0"
RBC_MID_BLUE = "5B9BD5"
RBC_WHITE = "FFFFFF"
STATUS_RED = "FF4444"
STATUS_YELLOW = "FFD966"
STATUS_GREEN = "70AD47"
STATUS_GREY = "D9D9D9"
FONT_MAIN = "Arial"

# Quarter-end months (calendar year quarters)
QUARTER_END_MONTHS = {3, 6, 9, 12}


# ── HELPERS ──────────────────────────────────────────────────────────────────
def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def header_font(size=10, bold=True, color=RBC_WHITE):
    return Font(name=FONT_MAIN, size=size, bold=bold, color=color)


def cell_font(size=9, bold=False, color="1F1F1F"):
    return Font(name=FONT_MAIN, size=size, bold=bold, color=color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_month_range(history=3, forward=5):
    """Return month-start dates: N history months + current month + N forward months."""
    today = date.today()
    y, m = today.year, today.month
    # Step back history months
    for _ in range(history):
        m -= 1
        if m < 1:
            m = 12
            y -= 1
    months = []
    total = history + 1 + forward
    for _ in range(total):
        months.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def gv_day_to_number(gv_str):
    if not gv_str:
        return None
    s = str(gv_str).strip()
    match = re.match(r"GV\s*(\d+)", s, re.IGNORECASE)
    if match:
        return int(match.group(1))
    match = re.match(r"(\d+)", s)
    if match:
        return int(match.group(1))
    return None


def nth_business_day(year, month, n):
    if n is None:
        return None
    d = date(year, month, 1)
    count = 0
    while True:
        if d.weekday() < 5:
            count += 1
            if count == n:
                return d
        d += timedelta(days=1)
        if d.month != month:
            return None


def is_quarterly(freq_str):
    """Return True if frequency indicates quarterly."""
    if not freq_str:
        return False
    f = freq_str.strip().lower()
    return "quarter" in f or f == "q" or f == "qtr"


def replace_katie(val):
    """Replace any occurrence of Katie (case-insensitive) with Brian."""
    if not val:
        return val
    return re.sub(r'\bKatie\b', 'Brian', str(val), flags=re.IGNORECASE)


def should_remove(name):
    """Return True if this deliverable should be removed."""
    if not name:
        return False
    return "aipi" in name.strip().lower()


# ── READ EXISTING DATA ───────────────────────────────────────────────────────
print("Loading workbook…")
wb = load_workbook(FILE_PATH)
if TAB_NAME not in wb.sheetnames:
    raise ValueError(f"Sheet '{TAB_NAME}' not found. Available: {wb.sheetnames}")
src = wb[TAB_NAME]

# The v1 rebuilt sheet has:
# Row 1 = title banner
# Row 2 = month span headers
# Row 3 = column headers: Name | Frequency | Timing | Owner | GV Day | …
# Row 4+ = data

# We read the static columns (A-E) from row 4 onward.
deliverables = []
for row_num in range(4, src.max_row + 1):
    name_val = src.cell(row=row_num, column=1).value
    if not name_val or not str(name_val).strip():
        continue
    name = str(name_val).strip()
    if name.upper() == "LEGEND":
        break

    freq = str(src.cell(row=row_num, column=2).value or "").strip()
    # Column 3 was "Timing" — we're dropping it, but read it in case GV Day is empty
    timing = str(src.cell(row=row_num, column=3).value or "").strip()
    owner = str(src.cell(row=row_num, column=4).value or "").strip()
    gv_raw = str(src.cell(row=row_num, column=5).value or "").strip()

    # If GV Day is empty but Timing has a value, use Timing as GV Day
    if not gv_raw and timing:
        gv_raw = timing

    # ── Apply Eric's changes ──
    # Remove AIPI
    if should_remove(name):
        print(f"  Removing: {name}")
        continue

    # Replace Katie → Brian
    owner = replace_katie(owner)

    # Clean up frequency: if it says both "Monthly" and "Quarterly", keep as-is
    # but we'll handle display logic via is_quarterly()

    deliverables.append({
        "name": name,
        "frequency": freq,
        "owner": owner,
        "gv_raw": gv_raw,
        "preparer": replace_katie(owner),
        "reviewed_by": "",
    })

print(f"Loaded {len(deliverables)} deliverables (after removals).")
if not deliverables:
    raise ValueError("No deliverables found after filtering.")


# ── DELETE OLD SHEET, CREATE FRESH ───────────────────────────────────────────
del wb[TAB_NAME]
ws = wb.create_sheet(TAB_NAME)

if "Done" in wb.sheetnames:
    sheet_order = wb.sheetnames
    current_idx = sheet_order.index(TAB_NAME)
    target_idx = sheet_order.index("Done")
    offset = target_idx - current_idx
    if offset > 0:
        wb.move_sheet(TAB_NAME, offset=offset - 1)


# ── COLUMN LAYOUT ────────────────────────────────────────────────────────────
# Static cols: Name | Frequency | Owner | GV Day (Timing removed per Eric)
STATIC_COLS = ["Name", "Frequency", "Owner", "GV Day"]
COL_WIDTHS = [35, 14, 14, 10]

# Per-month sub-columns
MONTH_COL_LABELS = ["Preparer", "Reviewed By", "Distributed Date", "Due Date"]
MONTH_COL_WIDTHS = [14, 14, 16, 12]

months = build_month_range(HISTORY_MONTHS, FORWARD_MONTHS)

STATIC_COUNT = len(STATIC_COLS)
MONTH_COLS = len(MONTH_COL_LABELS)
TOTAL_COLS = STATIC_COUNT + MONTH_COLS * len(months)

today = date.today()
current_month_start = date(today.year, today.month, 1)


# ── ROW 1: TITLE BANNER ─────────────────────────────────────────────────────
ws.row_dimensions[1].height = 28
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
title_cell = ws.cell(row=1, column=1, value="RECURRING DELIVERABLES TRACKER")
title_cell.font = Font(name=FONT_MAIN, size=14, bold=True, color=RBC_WHITE)
title_cell.fill = solid(RBC_NAVY)
title_cell.alignment = center()


# ── ROW 2: MONTH SPAN HEADERS ────────────────────────────────────────────────
ws.row_dimensions[2].height = 22

# Static area — blank navy fill
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=STATIC_COUNT)
for ci in range(1, STATIC_COUNT + 1):
    ws.cell(row=2, column=ci).fill = solid(RBC_NAVY)

for i, m in enumerate(months):
    start_col = STATIC_COUNT + 1 + i * MONTH_COLS
    end_col = start_col + MONTH_COLS - 1
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)

    # Label = reporting month (e.g., "NOV-25" means November reports)
    label = m.strftime("%b-%y").upper()

    # Mark history months vs current/future
    is_history = m < current_month_start

    mc = ws.cell(row=2, column=start_col, value=label)
    mc.font = header_font(size=10, bold=True, color=RBC_WHITE)
    mc.alignment = center()
    mc.border = thin_border("medium", RBC_WHITE)

    if m == current_month_start:
        # Highlight current month
        mc.fill = solid(RBC_MID_BLUE)
    elif is_history:
        # Slightly muted for history
        mc.fill = solid("3A5A8C")
    else:
        mc.fill = solid(RBC_BLUE if i % 2 == 0 else RBC_NAVY)


# ── ROW 3: COLUMN HEADERS ───────────────────────────────────────────────────
ws.row_dimensions[3].height = 36
for ci, label in enumerate(STATIC_COLS, start=1):
    c = ws.cell(row=3, column=ci, value=label)
    c.font = header_font(size=9)
    c.fill = solid(RBC_NAVY)
    c.alignment = center()
    c.border = thin_border("thin", RBC_WHITE)
    ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci - 1]

for i in range(len(months)):
    for j, label in enumerate(MONTH_COL_LABELS):
        col = STATIC_COUNT + 1 + i * MONTH_COLS + j
        c = ws.cell(row=3, column=col, value=label)
        c.font = header_font(size=8)
        c.fill = solid(RBC_MID_BLUE if i % 2 == 0 else RBC_BLUE)
        c.alignment = center()
        c.border = thin_border("thin", RBC_WHITE)
        ws.column_dimensions[get_column_letter(col)].width = MONTH_COL_WIDTHS[j]


# ── DATA ROWS ────────────────────────────────────────────────────────────────
DATA_START_ROW = 4

for ri, d in enumerate(deliverables):
    row_num = DATA_START_ROW + ri
    is_alt = ri % 2 == 1
    row_fill = solid(RBC_LIGHT_BLUE if is_alt else RBC_WHITE)
    ws.row_dimensions[row_num].height = 18

    quarterly = is_quarterly(d["frequency"])

    # Static columns (no more Timing column)
    static_vals = [d["name"], d["frequency"], d["owner"], d["gv_raw"]]
    for ci, val in enumerate(static_vals, start=1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.font = cell_font(size=9, bold=(ci == 1))
        c.fill = solid(RBC_NAVY) if ci == 1 else row_fill
        c.alignment = left() if ci == 1 else center()
        c.border = thin_border("thin")
        if ci == 1:
            c.font = Font(name=FONT_MAIN, size=9, bold=True, color=RBC_WHITE)

    # Month columns
    gv_num = gv_day_to_number(d["gv_raw"])

    for mi, m in enumerate(months):
        col_base = STATIC_COUNT + 1 + mi * MONTH_COLS
        due_date = nth_business_day(m.year, m.month, gv_num) if gv_num else None

        # If quarterly, only populate quarter-end months
        skip_month = quarterly and m.month not in QUARTER_END_MONTHS

        # Preparer
        c = ws.cell(row=row_num, column=col_base,
                     value="" if skip_month else d["preparer"])
        c.font = cell_font(8)
        c.fill = solid(STATUS_GREY) if skip_month else row_fill
        c.alignment = center()
        c.border = thin_border("thin")

        # Reviewed By
        c = ws.cell(row=row_num, column=col_base + 1, value="")
        c.font = cell_font(8)
        c.fill = solid(STATUS_GREY) if skip_month else row_fill
        c.alignment = center()
        c.border = thin_border("thin")

        # Distributed Date (user fills in)
        c = ws.cell(row=row_num, column=col_base + 2, value="")
        c.font = cell_font(8)
        c.fill = solid(STATUS_GREY) if skip_month else row_fill
        c.alignment = center()
        c.border = thin_border("thin")
        c.number_format = "MM/DD/YY"

        # Due Date
        if skip_month:
            c = ws.cell(row=row_num, column=col_base + 3, value="")
            c.font = Font(name=FONT_MAIN, size=8, color="888888")
            c.fill = solid(STATUS_GREY)
        elif due_date:
            c = ws.cell(row=row_num, column=col_base + 3, value=due_date)
            c.number_format = "M/D/YY"
            c.font = Font(name=FONT_MAIN, size=8, bold=True, color="1F1F1F")
            c.fill = row_fill
        else:
            label = "N/A" if not gv_num else "flexible"
            c = ws.cell(row=row_num, column=col_base + 3, value=label)
            c.font = Font(name=FONT_MAIN, size=8, bold=True, color="888888")
            c.fill = row_fill
        c.alignment = center()
        c.border = thin_border("thin")


# ── STATIC DUE-DATE CELL COLORING ────────────────────────────────────────────
for ri in range(len(deliverables)):
    row_num = DATA_START_ROW + ri
    d = deliverables[ri]
    gv_num = gv_day_to_number(d["gv_raw"])
    quarterly = is_quarterly(d["frequency"])

    for mi, m in enumerate(months):
        if quarterly and m.month not in QUARTER_END_MONTHS:
            continue
        col_base = STATIC_COUNT + 1 + mi * MONTH_COLS
        due_date = nth_business_day(m.year, m.month, gv_num) if gv_num else None
        if due_date is None:
            continue

        due_cell = ws.cell(row=row_num, column=col_base + 3)
        if today > due_date:
            due_cell.fill = solid(STATUS_RED)
            due_cell.font = Font(name=FONT_MAIN, size=8, bold=True, color=RBC_WHITE)
        elif (due_date - today).days <= 3:
            due_cell.fill = solid(STATUS_YELLOW)
            due_cell.font = Font(name=FONT_MAIN, size=8, bold=True, color="1F1F1F")
        else:
            due_cell.fill = solid(STATUS_GREEN)
            due_cell.font = Font(name=FONT_MAIN, size=8, bold=True, color=RBC_WHITE)


# ── EXCEL CONDITIONAL FORMATTING on Distributed Date cols ────────────────────
red_fill = PatternFill("solid", fgColor=STATUS_RED)
yellow_fill = PatternFill("solid", fgColor=STATUS_YELLOW)
green_fill = PatternFill("solid", fgColor=STATUS_GREEN)
red_font = Font(name=FONT_MAIN, bold=True, color=RBC_WHITE, size=8)
yellow_font = Font(name=FONT_MAIN, bold=True, color="1F1F1F", size=8)
green_font = Font(name=FONT_MAIN, bold=True, color=RBC_WHITE, size=8)

for mi, m in enumerate(months):
    col_base = STATIC_COUNT + 1 + mi * MONTH_COLS
    dist_col = col_base + 2
    due_col_num = col_base + 3
    dist_letter = get_column_letter(dist_col)
    due_letter = get_column_letter(due_col_num)

    row_start = DATA_START_ROW
    row_end = DATA_START_ROW + len(deliverables) - 1
    cell_range = f"{dist_letter}{row_start}:{dist_letter}{row_end}"
    first_dist = f"{dist_letter}{row_start}"
    first_due = f"{due_letter}{row_start}"

    # GREEN — distributed date filled in (highest priority)
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'NOT(ISBLANK({first_dist}))'],
        fill=green_fill, font=green_font, stopIfTrue=True
    ))

    # RED — blank AND today past the due date
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'AND(ISBLANK({first_dist}),ISNUMBER({first_due}),TODAY()>{first_due})'],
        fill=red_fill, font=red_font, stopIfTrue=True
    ))

    # YELLOW — blank AND due within 3 days
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'AND(ISBLANK({first_dist}),ISNUMBER({first_due}),TODAY()<={first_due},({first_due}-TODAY())<=3)'],
        fill=yellow_fill, font=yellow_font, stopIfTrue=True
    ))


# ── STATUS LEGEND ────────────────────────────────────────────────────────────
legend_row = DATA_START_ROW + len(deliverables) + 2
ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=STATIC_COUNT)
lbl = ws.cell(row=legend_row, column=1, value="LEGEND")
lbl.font = Font(name=FONT_MAIN, size=9, bold=True, color=RBC_WHITE)
lbl.fill = solid(RBC_NAVY)
lbl.alignment = center()

legend_items = [
    (STATUS_GREEN, RBC_WHITE, "Distributed"),
    (STATUS_YELLOW, "1F1F1F", "Due in 3 days or fewer"),
    (STATUS_RED, RBC_WHITE, "Overdue — not distributed"),
    (STATUS_GREY, "1F1F1F", "N/A, quarterly skip, or flexible"),
]
for li, (bg, fg, text) in enumerate(legend_items):
    r = legend_row + 1 + li
    ws.row_dimensions[r].height = 15
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT_MAIN, size=8, bold=True, color=fg)
    c.fill = solid(bg)
    c.alignment = left()
    c.border = thin_border("thin")


# ── FREEZE PANES ─────────────────────────────────────────────────────────────
ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=STATIC_COUNT + 1)


# ── AUTO-FILTER ──────────────────────────────────────────────────────────────
last_col_letter = get_column_letter(TOTAL_COLS)
ws.auto_filter.ref = f"A3:{last_col_letter}3"


# ── PRINT SETTINGS ───────────────────────────────────────────────────────────
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.sheet_view.showGridLines = False


# ── SAVE ─────────────────────────────────────────────────────────────────────
print("Saving…")
wb.save(FILE_PATH)
print(f"Done. File saved to: {FILE_PATH}")
print()
print("Changes applied:")
print("  - Timing column removed (redundant with GV Day)")
print("  - History months included: Nov-25, Dec-25, Jan-26")
print("  - Month headers = reporting month (not distribution month)")
print("  - Quarterly items: non-quarter-end months greyed out")
print("  - Katie replaced with Brian in all owner/preparer fields")
print("  - AIPI removed from deliverables list")
