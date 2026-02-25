"""
rebuild_recurring_v3.py

Scans the current Recurring sheet as-is, then rebuilds incorporating:

1. Hidden rows → moved to "Archived" section at the bottom
1. No alternating row colors (breaks when hiding rows) — uniform white + borders
1. Preparer removed from monthly sub-columns (stays in static area only)
1. Monthly cols = Reviewed By | Distributed Date | Due Date
1. History months: Nov, Dec, Jan + current + 5 forward
1. Reporting-month labeling (Jan column = January reports)
1. Timing column removed (redundant with GV Day)
1. Quarterly items: non-quarter months greyed out
1. Katie → Brian everywhere
1. AIPI removed

Usage:
    pip install openpyxl
    python rebuild_recurring_v3.py
"""

import re
from datetime import date, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────────
FILE_PATH = r"C:\Users\iams395\Copy of Project Tracking.xlsx"
TAB_NAME = "Recurring"
HISTORY_MONTHS = 3
FORWARD_MONTHS = 5

# ── COLOR PALETTE ────────────────────────────────────────────────────────────
RBC_NAVY = "001F5B"
RBC_BLUE = "0051A5"
RBC_MID_BLUE = "5B9BD5"
RBC_WHITE = "FFFFFF"
ROW_BG = "FFFFFF"          # uniform white for all data rows
BORDER_COLOR = "C0C0C0"   # subtle gray grid lines
SECTION_BORDER = "808080"  # darker border for section separators
STATUS_RED = "FF4444"
STATUS_YELLOW = "FFD966"
STATUS_GREEN = "70AD47"
STATUS_GREY = "D9D9D9"
ARCHIVED_BG = "F2F2F2"    # light grey for archived rows
FONT_MAIN = "Arial"

QUARTER_END_MONTHS = {3, 6, 9, 12}

# Items to remove entirely
REMOVE_NAMES = {"aipi"}

# Items to archive (hidden rows + explicitly named)
ARCHIVE_NAMES = {
    "fee-based monthly email",
    "fee-based forecast",
    "detailed metrics",
    "credit summary report",
    "ops inter-platform charges",
    "ops-inter-platform charges",
}

# ── HELPERS ──────────────────────────────────────────────────────────────────

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(color=BORDER_COLOR, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def section_top_border():
    """Heavier top border for section breaks."""
    top = Side(style="medium", color=SECTION_BORDER)
    rest = Side(style="thin", color=BORDER_COLOR)
    return Border(left=rest, right=rest, top=top, bottom=rest)

def header_font(size=10, bold=True, color=RBC_WHITE):
    return Font(name=FONT_MAIN, size=size, bold=bold, color=color)

def cell_font(size=9, bold=False, color="1F1F1F"):
    return Font(name=FONT_MAIN, size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def build_month_range(history=3, forward=5):
    today = date.today()
    y, m = today.year, today.month
    for _ in range(history):
        m -= 1
        if m < 1:
            m = 12
            y -= 1
    months = []
    for _ in range(history + 1 + forward):
        months.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months

def gv_day_to_number(gv_str):
    if not gv_str:
        return None
    s = str(gv_str).strip()
    match = re.match(r"GV\s*(\d+)", s, re.IGNORECASE)
    if match:
        return int(match.group(1))
    match = re.match(r"(\d+)", s)
    if match:
        return int(match.group(1))
    return None

def nth_business_day(year, month, n):
    if n is None:
        return None
    d = date(year, month, 1)
    count = 0
    while True:
        if d.weekday() < 5:
            count += 1
            if count == n:
                return d
        d += timedelta(days=1)
        if d.month != month:
            return None

def is_quarterly(freq_str):
    if not freq_str:
        return False
    f = freq_str.strip().lower()
    return "quarter" in f or "fiscal quarter" in f or f in ("q", "qtr")

def replace_katie(val):
    if not val:
        return val
    return re.sub(r'\bKatie\b', 'Brian', str(val), flags=re.IGNORECASE)

def normalize(s):
    return s.strip().lower() if s else ""

# ── SCAN EXISTING SHEET ──────────────────────────────────────────────────────
print("Loading workbook and scanning Recurring sheet…")
wb = load_workbook(FILE_PATH)
if TAB_NAME not in wb.sheetnames:
    raise ValueError(f"Sheet '{TAB_NAME}' not found. Available: {wb.sheetnames}")

src = wb[TAB_NAME]

# Detect the column layout dynamically by scanning for the header row
# Look for a row where one cell says "Name" and another says "Frequency"
header_row_num = None
col_map = {}
for row_num in range(1, min(src.max_row + 1, 10)):
    vals = {ci: (src.cell(row=row_num, column=ci).value or "")
            for ci in range(1, src.max_column + 1)}
    # Check if this row contains "Name" and "Frequency"
    name_col = None
    freq_col = None
    for ci, v in vals.items():
        vl = str(v).strip().lower()
        if vl == "name":
            name_col = ci
        elif vl == "frequency":
            freq_col = ci
    if name_col and freq_col:
        header_row_num = row_num
        # Map all header columns
        for ci, v in vals.items():
            vl = str(v).strip().lower()
            if vl:
                col_map[vl] = ci
        break

if not header_row_num:
    raise ValueError("Could not find header row with 'Name' and 'Frequency'.")

print(f"  Header row: {header_row_num}")
print(f"  Columns found: {col_map}")

# Column indices
NAME_COL = col_map.get("name", 1)
FREQ_COL = col_map.get("frequency", 2)
TIME_COL = col_map.get("timing")
OWNER_COL = col_map.get("owner", 4)
GV_COL = col_map.get("gv day", 5)
PREP_COL = col_map.get("preparer")

# Read all data rows, tracking which are hidden
active_items = []
archived_items = []
hidden_rows = set()

# Detect hidden rows
for row_num in range(header_row_num + 1, src.max_row + 1):
    rd = src.row_dimensions.get(row_num)
    if rd and rd.hidden:
        hidden_rows.add(row_num)

print(f"  Hidden rows detected: {sorted(hidden_rows)}")

for row_num in range(header_row_num + 1, src.max_row + 1):
    name_val = src.cell(row=row_num, column=NAME_COL).value
    if not name_val or not str(name_val).strip():
        continue
    name = str(name_val).strip()

    # Stop at legend or other non-data markers
    if name.upper() in ("LEGEND", "ARCHIVED"):
        break

    freq = str(src.cell(row=row_num, column=FREQ_COL).value or "").strip()
    owner = str(src.cell(row=row_num, column=OWNER_COL).value or "").strip()

    gv_raw = ""
    if GV_COL:
        gv_raw = str(src.cell(row=row_num, column=GV_COL).value or "").strip()
    # Fall back to Timing if GV Day is empty
    if not gv_raw and TIME_COL:
        gv_raw = str(src.cell(row=row_num, column=TIME_COL).value or "").strip()

    preparer = owner
    if PREP_COL:
        prep_val = src.cell(row=row_num, column=PREP_COL).value
        if prep_val and str(prep_val).strip():
            preparer = str(prep_val).strip()

    # Apply Katie → Brian
    owner = replace_katie(owner)
    preparer = replace_katie(preparer)

    # Remove AIPI entirely
    if normalize(name) in REMOVE_NAMES:
        print(f"  Removing: {name}")
        continue

    item = {
        "name": name,
        "frequency": freq,
        "owner": owner,
        "gv_raw": gv_raw,
        "preparer": preparer,
    }

    # Determine if archived: hidden row OR name matches archive list
    is_hidden = row_num in hidden_rows
    is_archive_name = normalize(name) in ARCHIVE_NAMES

    if is_hidden or is_archive_name:
        archived_items.append(item)
        print(f"  Archiving: {name} ({'hidden' if is_hidden else 'name match'})")
    else:
        active_items.append(item)

print(f"\n  Active: {len(active_items)}, Archived: {len(archived_items)}")
if not active_items:
    raise ValueError("No active deliverables found.")

# ── DELETE OLD SHEET, CREATE FRESH ───────────────────────────────────────────
del wb[TAB_NAME]
ws = wb.create_sheet(TAB_NAME)

if "Done" in wb.sheetnames:
    sheet_order = wb.sheetnames
    current_idx = sheet_order.index(TAB_NAME)
    target_idx = sheet_order.index("Done")
    offset = target_idx - current_idx
    if offset > 0:
        wb.move_sheet(TAB_NAME, offset=offset - 1)

# ── COLUMN LAYOUT ─────────────────────────────────────────────────────────────
# Static: Name | Frequency | Owner | GV Day | Preparer
# Per month: Reviewed By | Distributed Date | Due Date (no per-month Preparer)
STATIC_COLS = ["Name", "Frequency", "Owner", "GV Day", "Preparer"]
COL_WIDTHS = [38, 16, 12, 10, 12]

MONTH_COL_LABELS = ["Reviewed By", "Distributed Date", "Due Date"]
MONTH_COL_WIDTHS = [14, 16, 12]

months = build_month_range(HISTORY_MONTHS, FORWARD_MONTHS)
STATIC_COUNT = len(STATIC_COLS)
MONTH_COLS = len(MONTH_COL_LABELS)
TOTAL_COLS = STATIC_COUNT + MONTH_COLS * len(months)
today = date.today()
current_month_start = date(today.year, today.month, 1)

# ── ROW 1: TITLE BANNER ─────────────────────────────────────────────────────
ws.row_dimensions[1].height = 28
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
c = ws.cell(row=1, column=1, value="RECURRING DELIVERABLES TRACKER")
c.font = Font(name=FONT_MAIN, size=14, bold=True, color=RBC_WHITE)
c.fill = solid(RBC_NAVY)
c.alignment = center()

# ── ROW 2: MONTH SPAN HEADERS ────────────────────────────────────────────────
ws.row_dimensions[2].height = 22

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=STATIC_COUNT)
for ci in range(1, STATIC_COUNT + 1):
    ws.cell(row=2, column=ci).fill = solid(RBC_NAVY)

for i, m in enumerate(months):
    start_col = STATIC_COUNT + 1 + i * MONTH_COLS
    end_col = start_col + MONTH_COLS - 1
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    label = m.strftime("%b-%y").upper()
    is_history = m < current_month_start

    mc = ws.cell(row=2, column=start_col, value=label)
    mc.font = header_font(size=10, bold=True, color=RBC_WHITE)
    mc.alignment = center()
    mc.border = make_border(RBC_WHITE, "medium")

    if m == current_month_start:
        mc.fill = solid(RBC_MID_BLUE)
    elif is_history:
        mc.fill = solid("3A5A8C")
    else:
        mc.fill = solid(RBC_BLUE if i % 2 == 0 else RBC_NAVY)

# ── ROW 3: COLUMN HEADERS ─────────────────────────────────────────────────────
ws.row_dimensions[3].height = 36

for ci, label in enumerate(STATIC_COLS, start=1):
    c = ws.cell(row=3, column=ci, value=label)
    c.font = header_font(size=9)
    c.fill = solid(RBC_NAVY)
    c.alignment = center()
    c.border = make_border(RBC_WHITE)
    ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci - 1]

for i in range(len(months)):
    for j, label in enumerate(MONTH_COL_LABELS):
        col = STATIC_COUNT + 1 + i * MONTH_COLS + j
        c = ws.cell(row=3, column=col, value=label)
        c.font = header_font(size=8)
        c.fill = solid(RBC_MID_BLUE if i % 2 == 0 else RBC_BLUE)
        c.alignment = center()
        c.border = make_border(RBC_WHITE)
        ws.column_dimensions[get_column_letter(col)].width = MONTH_COL_WIDTHS[j]

# ── HELPER: WRITE A DATA ROW ─────────────────────────────────────────────────
def write_data_row(ws, row_num, d, months, is_archived=False):
    bg = solid(ARCHIVED_BG if is_archived else ROW_BG)
    bdr = make_border()
    ws.row_dimensions[row_num].height = 18
    quarterly = is_quarterly(d["frequency"])

    # Static columns
    static_vals = [d["name"], d["frequency"], d["owner"], d["gv_raw"], d["preparer"]]
    for ci, val in enumerate(static_vals, start=1):
        c = ws.cell(row=row_num, column=ci, value=val)
        if ci == 1:
            c.font = Font(name=FONT_MAIN, size=9, bold=True, color=RBC_WHITE)
            c.fill = solid(RBC_NAVY)
            c.alignment = left()
        else:
            c.font = cell_font(9)
            c.fill = bg
            c.alignment = center()
        c.border = bdr

    # Month columns
    gv_num = gv_day_to_number(d["gv_raw"])

    for mi, m in enumerate(months):
        col_base = STATIC_COUNT + 1 + mi * MONTH_COLS
        due_date = nth_business_day(m.year, m.month, gv_num) if gv_num else None
        skip = quarterly and m.month not in QUARTER_END_MONTHS

        # Reviewed By
        c = ws.cell(row=row_num, column=col_base, value="")
        c.font = cell_font(8)
        c.fill = solid(STATUS_GREY) if skip else bg
        c.alignment = center()
        c.border = bdr

        # Distributed Date
        c = ws.cell(row=row_num, column=col_base + 1, value="")
        c.font = cell_font(8)
        c.fill = solid(STATUS_GREY) if skip else bg
        c.alignment = center()
        c.border = bdr
        c.number_format = "MM/DD/YY"

        # Due Date
        if skip:
            c = ws.cell(row=row_num, column=col_base + 2, value="")
            c.font = Font(name=FONT_MAIN, size=8, color="888888")
            c.fill = solid(STATUS_GREY)
        elif due_date:
            c = ws.cell(row=row_num, column=col_base + 2, value=due_date)
            c.number_format = "M/D/YY"
            c.font = Font(name=FONT_MAIN, size=8, bold=True, color="1F1F1F")
            c.fill = bg
        else:
            lbl = "N/A" if not gv_num else "flexible"
            c = ws.cell(row=row_num, column=col_base + 2, value=lbl)
            c.font = Font(name=FONT_MAIN, size=8, bold=True, color="888888")
            c.fill = bg

        c.alignment = center()
        c.border = bdr

# ── WRITE ACTIVE ROWS ─────────────────────────────────────────────────────────
DATA_START_ROW = 4

for ri, d in enumerate(active_items):
    write_data_row(ws, DATA_START_ROW + ri, d, months, is_archived=False)

ACTIVE_END_ROW = DATA_START_ROW + len(active_items) - 1

# ── COLOR DUE DATE CELLS (active rows) ────────────────────────────────────────
for ri, d in enumerate(active_items):
    row_num = DATA_START_ROW + ri
    gv_num = gv_day_to_number(d["gv_raw"])
    quarterly = is_quarterly(d["frequency"])

    for mi, m in enumerate(months):
        if quarterly and m.month not in QUARTER_END_MONTHS:
            continue
        col_base = STATIC_COUNT + 1 + mi * MONTH_COLS
        due_date = nth_business_day(m.year, m.month, gv_num) if gv_num else None
        if not due_date:
            continue

        due_cell = ws.cell(row=row_num, column=col_base + 2)
        if today > due_date:
            due_cell.fill = solid(STATUS_RED)
            due_cell.font = Font(name=FONT_MAIN, size=8, bold=True, color=RBC_WHITE)
        elif (due_date - today).days <= 3:
            due_cell.fill = solid(STATUS_YELLOW)
            due_cell.font = Font(name=FONT_MAIN, size=8, bold=True, color="1F1F1F")
        else:
            due_cell.fill = solid(STATUS_GREEN)
            due_cell.font = Font(name=FONT_MAIN, size=8, bold=True, color=RBC_WHITE)

# ── CONDITIONAL FORMATTING on Distributed Date cols (active rows only) ────────
red_fill = PatternFill("solid", fgColor=STATUS_RED)
yellow_fill = PatternFill("solid", fgColor=STATUS_YELLOW)
green_fill = PatternFill("solid", fgColor=STATUS_GREEN)

red_font = Font(name=FONT_MAIN, bold=True, color=RBC_WHITE, size=8)
yellow_font = Font(name=FONT_MAIN, bold=True, color="1F1F1F", size=8)
green_font = Font(name=FONT_MAIN, bold=True, color=RBC_WHITE, size=8)

for mi in range(len(months)):
    col_base = STATIC_COUNT + 1 + mi * MONTH_COLS
    dist_col = col_base + 1       # Distributed Date is now column index +1
    due_col_num = col_base + 2     # Due Date is +2
    dist_letter = get_column_letter(dist_col)
    due_letter = get_column_letter(due_col_num)
    row_start = DATA_START_ROW
    row_end = ACTIVE_END_ROW
    cell_range = f"{dist_letter}{row_start}:{dist_letter}{row_end}"
    first_dist = f"{dist_letter}{row_start}"
    first_due = f"{due_letter}{row_start}"

    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'NOT(ISBLANK({first_dist}))'],
        fill=green_fill, font=green_font, stopIfTrue=True
    ))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'AND(ISBLANK({first_dist}),ISNUMBER({first_due}),TODAY()>{first_due})'],
        fill=red_fill, font=red_font, stopIfTrue=True
    ))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'AND(ISBLANK({first_dist}),ISNUMBER({first_due}),TODAY()<={first_due},({first_due}-TODAY())<=3)'],
        fill=yellow_fill, font=yellow_font, stopIfTrue=True
    ))

# ── ARCHIVED SECTION ──────────────────────────────────────────────────────────
if archived_items:
    archive_header_row = ACTIVE_END_ROW + 3

    # Section header
    ws.merge_cells(start_row=archive_header_row, start_column=1,
                   end_row=archive_header_row, end_column=TOTAL_COLS)
    c = ws.cell(row=archive_header_row, column=1, value="ARCHIVED DELIVERABLES")
    c.font = Font(name=FONT_MAIN, size=11, bold=True, color=RBC_WHITE)
    c.fill = solid("4A4A4A")
    c.alignment = center()
    ws.row_dimensions[archive_header_row].height = 24

    # Sub-header row
    sub_row = archive_header_row + 1
    ws.row_dimensions[sub_row].height = 28
    for ci, label in enumerate(STATIC_COLS, start=1):
        c = ws.cell(row=sub_row, column=ci, value=label)
        c.font = header_font(size=8, color=RBC_WHITE)
        c.fill = solid("666666")
        c.alignment = center()
        c.border = make_border("999999")

    for i in range(len(months)):
        for j, label in enumerate(MONTH_COL_LABELS):
            col = STATIC_COUNT + 1 + i * MONTH_COLS + j
            c = ws.cell(row=sub_row, column=col, value=label)
            c.font = header_font(size=7, color=RBC_WHITE)
            c.fill = solid("666666")
            c.alignment = center()
            c.border = make_border("999999")

    # Archived data rows
    arch_start = sub_row + 1
    for ai, d in enumerate(archived_items):
        write_data_row(ws, arch_start + ai, d, months, is_archived=True)

    LAST_DATA_ROW = arch_start + len(archived_items) - 1
else:
    LAST_DATA_ROW = ACTIVE_END_ROW

# ── LEGEND ────────────────────────────────────────────────────────────────────
legend_row = LAST_DATA_ROW + 3
ws.merge_cells(start_row=legend_row, start_column=1,
               end_row=legend_row, end_column=STATIC_COUNT)
lbl = ws.cell(row=legend_row, column=1, value="LEGEND")
lbl.font = Font(name=FONT_MAIN, size=9, bold=True, color=RBC_WHITE)
lbl.fill = solid(RBC_NAVY)
lbl.alignment = center()

legend_items = [
    (STATUS_GREEN, RBC_WHITE, "Distributed"),
    (STATUS_YELLOW, "1F1F1F", "Due in 3 days or fewer"),
    (STATUS_RED, RBC_WHITE, "Overdue — not distributed"),
    (STATUS_GREY, "1F1F1F", "N/A, quarterly skip, or flexible"),
    (ARCHIVED_BG, "1F1F1F", "Archived — no longer active"),
]
for li, (bg, fg, text) in enumerate(legend_items):
    r = legend_row + 1 + li
    ws.row_dimensions[r].height = 15
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT_MAIN, size=8, bold=True, color=fg)
    c.fill = solid(bg)
    c.alignment = left()
    c.border = make_border()

# ── FREEZE PANES ──────────────────────────────────────────────────────────────
ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=STATIC_COUNT + 1)

# ── AUTO-FILTER (active section only) ─────────────────────────────────────────
last_col_letter = get_column_letter(TOTAL_COLS)
ws.auto_filter.ref = f"A3:{last_col_letter}{ACTIVE_END_ROW}"

# ── PRINT SETTINGS ────────────────────────────────────────────────────────────
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.sheet_view.showGridLines = False

# ── SAVE ──────────────────────────────────────────────────────────────────────
print("\nSaving…")
wb.save(FILE_PATH)
print(f"Done. File saved to: {FILE_PATH}")
print()
print("Summary of changes:")
print(f"  - {len(active_items)} active deliverables")
print(f"  - {len(archived_items)} archived deliverables (moved to bottom section)")
print("  - Timing column removed")
print("  - Per-month Preparer removed (Preparer stays in static cols)")
print("  - Alternating row colors removed (uniform white + borders)")
print("  - Quarterly items: non-quarter months greyed out")
print("  - Katie replaced with Brian everywhere")
print("  - AIPI removed")
print("  - History months: Nov-25, Dec-25, Jan-26 included")
